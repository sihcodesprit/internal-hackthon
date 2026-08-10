import qrcode
import json
import io
import base64
import uuid
from datetime import timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.files.base import ContentFile
from django.db.models import Count, Q
from django.views.decorators.csrf import csrf_exempt

from .models import Class, Subject, AttendanceSession, AttendanceRecord
from .forms import ClassForm, AttendanceSessionForm
from accounts.models import CustomUser
from timetable.models import TimeSlot


def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.user.is_teacher():
        return teacher_dashboard(request)
    elif request.user.is_student():
        return student_dashboard(request)
    else:
        return redirect('admin:index')


def teacher_dashboard(request):
    teacher = request.user
    classes = Class.objects.filter(teacher=teacher).select_related('subject')
    recent_sessions = AttendanceSession.objects.filter(
        teacher=teacher
    ).select_related('class_obj').order_by('-created_at')[:5]

    today = timezone.now().date()
    today_sessions = AttendanceSession.objects.filter(teacher=teacher, date=today)
    active_sessions = today_sessions.filter(status='active')

    total_students = sum(c.students.count() for c in classes)

    today_day = today.strftime('%A')[:3].upper()
    days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
    day_labels = {'MON': 'Monday', 'TUE': 'Tuesday', 'WED': 'Wednesday',
                  'THU': 'Thursday', 'FRI': 'Friday', 'SAT': 'Saturday'}
    week_slots = TimeSlot.objects.filter(
        class_obj__teacher=teacher
    ).select_related('class_obj__subject').order_by('day', 'start_time')
    week_timetable = {day: [] for day in days}
    for slot in week_slots:
        week_timetable[slot.day].append(slot)
    today_slots = week_timetable.get(today_day, [])

    context = {
        'classes': classes,
        'recent_sessions': recent_sessions,
        'today_sessions': today_sessions,
        'active_sessions': active_sessions,
        'total_students': total_students,
        'total_classes': classes.count(),
        'today': today,
        'today_day': today_day,
        'days': days,
        'day_labels': day_labels,
        'week_timetable': week_timetable,
        'today_slots': today_slots,
    }
    return render(request, 'attendance/teacher_dashboard.html', context)


def student_dashboard(request):
    student = request.user
    enrolled_classes = student.enrolled_classes.select_related('subject', 'teacher').all()
    records = AttendanceRecord.objects.filter(student=student).select_related('session__class_obj')

    # Calculate attendance percentage per subject
    attendance_stats = []
    for cls in enrolled_classes:
        total_sessions = AttendanceSession.objects.filter(class_obj=cls).count()
        present_count = AttendanceRecord.objects.filter(
            student=student, session__class_obj=cls, status='present'
        ).count()
        percentage = round((present_count / total_sessions * 100) if total_sessions > 0 else 0, 1)
        attendance_stats.append({
            'class': cls,
            'total': total_sessions,
            'present': present_count,
            'percentage': percentage,
            'status': 'safe' if percentage >= 75 else ('warning' if percentage >= 60 else 'danger')
        })

    today = timezone.now().date()
    today_day = today.strftime('%A')[:3].upper()
    recent_records = records.order_by('-marked_at')[:10]

    today_slots = TimeSlot.objects.filter(
        department=student.department,
        year=student.year_of_study,
        section=student.section,
        day=today_day
    ).select_related('class_obj__subject', 'class_obj__teacher').order_by('start_time')

    # Active sessions student can mark
    active_sessions = AttendanceSession.objects.filter(
        class_obj__in=enrolled_classes,
        status='active'
    ).exclude(records__student=student)

    context = {
        'enrolled_classes': enrolled_classes,
        'attendance_stats': attendance_stats,
        'recent_records': recent_records,
        'active_sessions': active_sessions,
        'today': today,
        'today_day': today_day,
        'today_slots': today_slots,
    }
    return render(request, 'attendance/student_dashboard.html', context)


@login_required
def class_list(request):
    if request.user.is_teacher():
        classes = Class.objects.filter(teacher=request.user).select_related('subject')
    else:
        classes = request.user.enrolled_classes.select_related('subject', 'teacher')
    return render(request, 'attendance/class_list.html', {'classes': classes})


@login_required
def class_create(request):
    if not request.user.is_teacher():
        messages.error(request, 'Only teachers can create classes.')
        return redirect('dashboard')
    if request.method == 'POST':
        form = ClassForm(request.POST)
        if form.is_valid():
            cls = form.save(commit=False)
            cls.teacher = request.user
            cls.save()
            form.save_m2m()
            messages.success(request, f'Class "{cls.name}" created successfully!')
            return redirect('class_list')
    else:
        form = ClassForm()
    return render(request, 'attendance/class_form.html', {'form': form, 'title': 'Create Class'})


@login_required
def class_detail(request, pk):
    cls = get_object_or_404(Class, pk=pk)
    sessions = AttendanceSession.objects.filter(class_obj=cls).order_by('-created_at')

    student_stats = []
    for student in cls.students.all():
        total = sessions.count()
        present = AttendanceRecord.objects.filter(
            student=student, session__class_obj=cls, status='present'
        ).count()
        pct = round((present / total * 100) if total > 0 else 0, 1)
        student_stats.append({
            'student': student,
            'total': total,
            'present': present,
            'percentage': pct
        })

    context = {
        'class': cls,
        'sessions': sessions[:20],
        'student_stats': student_stats,
    }
    return render(request, 'attendance/class_detail.html', context)


@login_required
def generate_qr(request, class_id):
    if not request.user.is_teacher():
        messages.error(request, 'Only teachers can generate QR codes.')
        return redirect('dashboard')

    cls = get_object_or_404(Class, pk=class_id, teacher=request.user)
    duration = int(request.POST.get('duration', 15))  # minutes

    # Close any existing active sessions
    AttendanceSession.objects.filter(class_obj=cls, status='active').update(status='closed')

    session = AttendanceSession.objects.create(
        class_obj=cls,
        teacher=request.user,
        date=timezone.now().date(),
        start_time=timezone.now().time(),
        status='active',
        expires_at=timezone.now() + timedelta(minutes=duration),
    )

    # Generate QR code data
    qr_data = json.dumps({
        'session_id': str(session.session_id),
        'class': cls.name,
        'subject': cls.subject.code,
        'timestamp': timezone.now().isoformat(),
    })
    session.qr_data = qr_data
    session.save()

    # Generate QR image
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')

    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    # Save QR image
    session.qr_code.save(f'qr_{session.session_id}.png', ContentFile(buffer.getvalue()), save=True)

    messages.success(request, f'QR code generated! Valid for {duration} minutes.')
    return render(request, 'attendance/qr_display.html', {
        'session': session,
        'qr_base64': qr_base64,
        'class': cls,
        'expires_at': session.expires_at,
        'duration': duration,
    })


@login_required
def mark_attendance(request):
    if not request.user.is_student():
        messages.error(request, 'Only students can mark attendance.')
        return redirect('dashboard')

    if request.method == 'POST':
        session_id = request.POST.get('session_id')
        try:
            session_data = json.loads(session_id)
            sid = session_data.get('session_id')
        except (json.JSONDecodeError, TypeError):
            sid = session_id

        try:
            session = AttendanceSession.objects.get(session_id=sid)
            if not session.is_active():
                return JsonResponse({'success': False, 'message': 'Session has expired or is closed.'})

            if not request.user.enrolled_classes.filter(pk=session.class_obj.pk).exists():
                return JsonResponse({'success': False, 'message': 'You are not enrolled in this class.'})

            record, created = AttendanceRecord.objects.get_or_create(
                session=session,
                student=request.user,
                defaults={
                    'status': 'present',
                    'ip_address': request.META.get('REMOTE_ADDR'),
                    'device_info': request.META.get('HTTP_USER_AGENT', '')[:500],
                }
            )
            if created:
                return JsonResponse({'success': True, 'message': f'Attendance marked for {session.class_obj.subject.name}!'})
            else:
                return JsonResponse({'success': False, 'message': 'Attendance already marked for this session.'})
        except AttendanceSession.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid QR code or session not found.'})

    return render(request, 'attendance/mark_attendance.html')


@login_required
def session_detail(request, session_id):
    session = get_object_or_404(AttendanceSession, session_id=session_id)
    records = AttendanceRecord.objects.filter(session=session).select_related('student')
    total_students = session.class_obj.students.count()
    present_count = records.filter(status='present').count()
    absent_count = total_students - present_count

    # Mark absent students
    absent_students = session.class_obj.students.exclude(
        attendance_records__session=session
    )

    context = {
        'session': session,
        'records': records,
        'absent_students': absent_students,
        'present_count': present_count,
        'absent_count': absent_count,
        'total_students': total_students,
        'attendance_pct': round((present_count / total_students * 100) if total_students > 0 else 0, 1),
    }
    return render(request, 'attendance/session_detail.html', context)


@login_required
def close_session(request, session_id):
    session = get_object_or_404(AttendanceSession, session_id=session_id, teacher=request.user)
    session.status = 'closed'
    session.save()
    messages.success(request, 'Session closed successfully.')
    return redirect('session_detail', session_id=session_id)


@login_required
def attendance_history(request):
    if request.user.is_teacher():
        sessions = AttendanceSession.objects.filter(teacher=request.user).select_related('class_obj').order_by('-date')
    else:
        records = AttendanceRecord.objects.filter(student=request.user).select_related('session__class_obj')
        return render(request, 'attendance/student_history.html', {'records': records})
    return render(request, 'attendance/teacher_history.html', {'sessions': sessions})


@login_required
def export_attendance_excel(request, session_id):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    session = get_object_or_404(AttendanceSession, session_id=session_id)
    records = AttendanceRecord.objects.filter(session=session).select_related('student')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header styling
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ['#', 'Roll Number', 'Name', 'Email', 'Status', 'Marked At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=record.student.roll_number)
        ws.cell(row=row_idx, column=3, value=record.student.get_full_name())
        ws.cell(row=row_idx, column=4, value=record.student.email)
        ws.cell(row=row_idx, column=5, value=record.status.capitalize())
        ws.cell(row=row_idx, column=6, value=str(record.marked_at))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_{session.session_id}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_attendance_pdf(request, session_id):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    session = get_object_or_404(AttendanceSession, session_id=session_id)
    records = AttendanceRecord.objects.filter(session=session).select_related('student')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{session.session_id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor('#1A1A2E'))
    elements.append(Paragraph(f"Attendance Report", title_style))
    elements.append(Paragraph(f"Class: {session.class_obj.name} | Date: {session.date}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))

    data = [['#', 'Roll No', 'Name', 'Status', 'Time']]
    for i, record in enumerate(records, 1):
        data.append([
            str(i),
            record.student.roll_number,
            record.student.get_full_name(),
            record.status.capitalize(),
            record.marked_at.strftime('%H:%M:%S')
        ])

    table = Table(data, colWidths=[0.4*inch, 1*inch, 2.5*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A2E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4FF')]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)
    return response
